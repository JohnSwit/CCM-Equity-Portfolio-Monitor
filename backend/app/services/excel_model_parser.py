"""
Excel Model Parser - Extracts data from analyst Excel models' API tabs

This module reads Excel files (local, OneDrive-synced, or via OneDrive share
links) and extracts standardized metrics from the "API" tab.

Expected API Tab Layout:
- E11: IRR (3-year)
- E13: CCM Fair Value
- E14: Street Price Target

Revenue/EBITDA/EPS/FCF estimates laid out in a grid format.
"""
import os
import tempfile
from typing import Dict, Any, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


def parse_excel_model(file_path: str) -> Dict[str, Any]:
    """
    Parse an Excel model file and extract data from the API tab.

    Args:
        file_path: Path to the Excel file (local or OneDrive-synced path)

    Returns:
        Dictionary with extracted model data

    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the API tab is missing or malformed
    """
    # Normalize the path (handle OneDrive paths)
    normalized_path = normalize_path(file_path)

    if not os.path.exists(normalized_path):
        raise FileNotFoundError(f"Excel file not found: {normalized_path}")

    return _parse_workbook(normalized_path)


def parse_excel_model_from_url(share_url: str) -> Dict[str, Any]:
    """
    Download an Excel model from a OneDrive/SharePoint share link and parse it.
    Uses the Microsoft Graph API with Azure AD client credentials for authentication.

    Args:
        share_url: OneDrive or SharePoint sharing URL

    Returns:
        Dictionary with extracted model data

    Raises:
        ConnectionError: If the file cannot be downloaded
        ValueError: If the API tab is missing or malformed
    """
    import httpx
    from app.core.config import settings

    logger.info("Downloading Excel model from share link via Graph API...")

    tmp_path = None
    try:
        # Get OAuth token for Microsoft Graph API
        access_token = _get_graph_access_token(settings)

        # Encode the sharing URL for Graph API
        encoded_url = _encode_sharing_url(share_url)

        # Download via Graph API: /shares/{encoded}/driveItem/content
        graph_url = f"https://graph.microsoft.com/v1.0/shares/{encoded_url}/driveItem/content"

        with httpx.Client(follow_redirects=True, timeout=60.0) as client:
            response = client.get(
                graph_url,
                headers={"Authorization": f"Bearer {access_token}"},
            )
            response.raise_for_status()

        # Verify we got an Excel file (not an error page)
        content_type = response.headers.get('content-type', '')
        if 'text/html' in content_type or 'application/json' in content_type:
            raise ConnectionError(
                f"Graph API returned unexpected content type: {content_type}. "
                "Check that the share link is valid and the app has Files.Read.All permission."
            )

        # Save to temp file
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        with os.fdopen(tmp_fd, 'wb') as f:
            f.write(response.content)

        logger.info(f"Downloaded {len(response.content)} bytes via Graph API")

        return _parse_workbook(tmp_path)

    except httpx.HTTPStatusError as e:
        error_body = e.response.text[:300] if e.response else ""
        if e.response.status_code == 401:
            raise ConnectionError(
                "Graph API authentication failed (401). "
                "Check AZURE_TENANT_ID, AZURE_CLIENT_ID, and AZURE_CLIENT_SECRET."
            )
        elif e.response.status_code == 403:
            raise ConnectionError(
                "Graph API access denied (403). "
                "Ensure the Azure AD app has 'Files.Read.All' application permission with admin consent."
            )
        raise ConnectionError(f"Failed to download model: HTTP {e.response.status_code} - {error_body}")
    except httpx.RequestError as e:
        raise ConnectionError(f"Failed to download model: {str(e)}")
    finally:
        # Clean up temp file
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


def _get_graph_access_token(settings) -> str:
    """
    Get an OAuth2 access token for Microsoft Graph API using client credentials flow.
    Requires AZURE_TENANT_ID, AZURE_CLIENT_ID, and AZURE_CLIENT_SECRET in settings.
    """
    import httpx

    if not settings.AZURE_TENANT_ID or not settings.AZURE_CLIENT_ID or not settings.AZURE_CLIENT_SECRET:
        raise ConnectionError(
            "Azure AD credentials not configured. "
            "Set AZURE_TENANT_ID, AZURE_CLIENT_ID, and AZURE_CLIENT_SECRET in .env "
            "to enable OneDrive/SharePoint model downloads."
        )

    token_url = f"https://login.microsoftonline.com/{settings.AZURE_TENANT_ID}/oauth2/v2.0/token"

    with httpx.Client(timeout=30.0) as client:
        response = client.post(
            token_url,
            data={
                "grant_type": "client_credentials",
                "client_id": settings.AZURE_CLIENT_ID,
                "client_secret": settings.AZURE_CLIENT_SECRET,
                "scope": "https://graph.microsoft.com/.default",
            },
        )
        response.raise_for_status()

    token_data = response.json()
    access_token = token_data.get("access_token")
    if not access_token:
        raise ConnectionError("Failed to obtain Graph API access token")

    return access_token


def _encode_sharing_url(share_url: str) -> str:
    """
    Encode a OneDrive/SharePoint sharing URL for use with the Graph API /shares endpoint.

    Microsoft Graph requires sharing URLs to be encoded as:
    1. Base64 encode the URL
    2. Replace '/' with '_', '+' with '-'
    3. Trim trailing '='
    4. Prepend 'u!'

    See: https://learn.microsoft.com/en-us/graph/api/shares-get
    """
    import base64

    url = share_url.strip()
    # Base64 encode
    encoded = base64.b64encode(url.encode('utf-8')).decode('utf-8')
    # Make URL-safe: replace + with -, / with _, trim =
    encoded = encoded.replace('+', '-').replace('/', '_').rstrip('=')
    # Prepend u!
    return f"u!{encoded}"


def _parse_workbook(file_path: str) -> Dict[str, Any]:
    """
    Parse an Excel workbook at the given local path and extract API tab data.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Error opening Excel file: {str(e)}")

    # Find the API tab
    api_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name.upper() == "API":
            api_sheet = wb[sheet_name]
            break

    if not api_sheet:
        raise ValueError("API tab not found in Excel model")

    # Extract data
    data = extract_api_tab_data(api_sheet)
    data['data_as_of'] = datetime.utcnow()

    wb.close()
    return data


def normalize_path(file_path: str) -> str:
    """
    Normalize file path, handling OneDrive paths and environment variables.

    Supports:
    - Local absolute paths
    - OneDrive paths (resolves to local sync folder)
    - Environment variable expansion
    - User home directory expansion (~)
    """
    # Expand environment variables
    path = os.path.expandvars(file_path)

    # Expand user home directory
    path = os.path.expanduser(path)

    # Handle OneDrive paths
    # Common OneDrive locations on Windows
    onedrive_roots = [
        os.path.expandvars(r"%USERPROFILE%\OneDrive"),
        os.path.expandvars(r"%USERPROFILE%\OneDrive - CCM"),
        os.path.expanduser("~/OneDrive"),
        os.path.expanduser("~/OneDrive - CCM"),
    ]

    # If path starts with OneDrive:// or similar, try to resolve it
    if path.lower().startswith("onedrive://"):
        relative_path = path[11:]  # Remove "onedrive://"
        for root in onedrive_roots:
            candidate = os.path.join(root, relative_path)
            if os.path.exists(candidate):
                return candidate
        # Return first candidate if none exist (let caller handle error)
        return os.path.join(onedrive_roots[0], relative_path)

    return path


def safe_float(value) -> Optional[float]:
    """Safely convert a cell value to float"""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def extract_api_tab_data(sheet) -> Dict[str, Any]:
    """
    Extract all required data from the API tab.

    Expected layout (adjust cell references based on your actual template):
    - E11: IRR 3-year
    - E13: CCM Fair Value
    - E14: Street Price Target

    For estimates, we expect a grid layout like:
                    -1YR    1YR     2YR     3YR
    Revenue CCM     B20     C20     D20     E20
    Revenue Street  B21     C21     D21     E21
    EBITDA CCM      B23     C23     D23     E23
    EBITDA Street   B24     C24     D24     E24
    EPS CCM         B26     C26     D26     E26
    EPS Street      B27     C27     D27     E27
    FCF CCM         B29     C29     D29     E29
    FCF Street      B30     C30     D30     E30

    Note: Adjust these cell references to match your actual template!
    """
    data = {}

    # Extract key metrics from fixed cells
    data['irr_3yr'] = safe_float(sheet['E11'].value)
    data['ccm_fair_value'] = safe_float(sheet['E13'].value)
    data['street_price_target'] = safe_float(sheet['E14'].value)

    # Define the grid layout for estimates
    # Format: (metric_name, ccm_row, street_row)
    # Columns: B=-1YR, C=1YR, D=2YR, E=3YR
    estimate_layout = [
        ('revenue', 20, 21),
        ('ebitda', 23, 24),
        ('eps', 26, 27),
        ('fcf', 29, 30),
    ]

    period_columns = {
        'minus1yr': 'B',
        '1yr': 'C',
        '2yr': 'D',
        '3yr': 'E',
    }

    for metric, ccm_row, street_row in estimate_layout:
        for period, col in period_columns.items():
            # CCM estimate
            ccm_cell = f"{col}{ccm_row}"
            data[f"{metric}_ccm_{period}"] = safe_float(sheet[ccm_cell].value)

            # Street estimate
            street_cell = f"{col}{street_row}"
            data[f"{metric}_street_{period}"] = safe_float(sheet[street_cell].value)

    return data


def create_api_tab_template() -> str:
    """
    Returns documentation for the expected API tab format.
    Users can use this to set up their Excel models correctly.
    """
    template = """
    API Tab Template for CCM Equity Models
    =======================================

    The API tab should contain the following data:

    VALUATION METRICS (Fixed Cells):
    --------------------------------
    Cell E11: 3-Year IRR (decimal, e.g., 0.15 for 15%)
    Cell E13: CCM Fair Value (price per share)
    Cell E14: Street Price Target (price per share)

    ESTIMATES GRID:
    ---------------
    Columns: B = -1YR, C = 1YR, D = 2YR, E = 3YR

    Revenue:
        Row 20: CCM Estimates
        Row 21: Street Estimates

    EBITDA:
        Row 23: CCM Estimates
        Row 24: Street Estimates

    EPS:
        Row 26: CCM Estimates
        Row 27: Street Estimates

    FCF (Free Cash Flow):
        Row 29: CCM Estimates
        Row 30: Street Estimates

    NOTES:
    ------
    - All values should be numbers (not text)
    - Revenue, EBITDA, FCF typically in millions
    - EPS in dollars per share
    - IRR as decimal (0.15 = 15%)
    - Fair values and price targets in dollars per share
    """
    return template


def validate_model_path(file_path: str) -> Dict[str, Any]:
    """
    Validate that a model path is accessible and has the expected format.

    Returns:
        Dictionary with validation results
    """
    result = {
        "valid": False,
        "exists": False,
        "has_api_tab": False,
        "errors": []
    }

    try:
        normalized_path = normalize_path(file_path)
        result["normalized_path"] = normalized_path

        if not os.path.exists(normalized_path):
            result["errors"].append("File not found")
            return result

        result["exists"] = True

        import openpyxl
        wb = openpyxl.load_workbook(normalized_path, data_only=True)

        # Check for API tab
        has_api = any(s.upper() == "API" for s in wb.sheetnames)
        result["has_api_tab"] = has_api
        result["sheets"] = wb.sheetnames

        if not has_api:
            result["errors"].append("API tab not found")

        wb.close()

        if has_api:
            result["valid"] = True

    except ImportError:
        result["errors"].append("openpyxl not installed")
    except Exception as e:
        result["errors"].append(str(e))

    return result
